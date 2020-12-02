from openpyxl import load_workbook

wb = load_workbook('/home/ashb245/python_training/Meal_Calculator.xlsx', data_only=True)
ws = wb['Sheet1']

print(ws['G7'].value)
print(ws['G7'].coordinate)

print(ws['G7'])

cell_range = ws['G10':'H17']

print(str(cell_range) + "\n\n\n")

for cell in cell_range:
    print(f"{cell[0].value} {cell[1].value}")